import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_DIR = os.path.join(os.path.dirname(__file__), '..', 'exports')
os.makedirs(EXCEL_DIR, exist_ok=True)

def generate_orders_excel(orders, date_str=None, is_master=False):
    import json

    if date_str:
        try:
            filter_date = date.fromisoformat(date_str)
            title = f"Orders for {filter_date.strftime('%d %b %Y')}"
        except:
            title = "All Orders"
    else:
        title = "All Orders"

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Styles
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    title_font = Font(name='Calibri', bold=True, size=16, color="7C3AED")
    subtitle_font = Font(name='Calibri', size=11, color="5C4F72")
    status_fills = {
        'pending': PatternFill(start_color="FEF3C7", fill_type="solid"),
        'approved': PatternFill(start_color="D1FAE5", fill_type="solid"),
        'rejected': PatternFill(start_color="FEE2E2", fill_type="solid"),
        'confirmed': PatternFill(start_color="D1FAE5", fill_type="solid"),
    }
    border = Border(
        left=Side(style='thin', color='E5E0F0'),
        right=Side(style='thin', color='E5E0F0'),
        top=Side(style='thin', color='E5E0F0'),
        bottom=Side(style='thin', color='E5E0F0')
    )

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = f"MenuMatrix - {title}"
    ws['A1'].font = title_font
    ws.merge_cells('A2:K2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws['A2'].font = subtitle_font

    # Summary
    total = len(orders)
    confirmed = sum(1 for o in orders if o.status == 'confirmed')
    pending = sum(1 for o in orders if o.status == 'pending')
    revenue = sum(o.grand_total for o in orders if o.status == 'confirmed')

    ws.merge_cells('A4:B4'); ws['A4'] = f"Total Orders: {total}"; ws['A4'].font = Font(bold=True, size=11)
    ws.merge_cells('C4:D4'); ws['C4'] = f"Confirmed: {confirmed}"; ws['C4'].font = Font(bold=True, color="10B981")
    ws.merge_cells('E4:F4'); ws['E4'] = f"Pending: {pending}"; ws['E4'].font = Font(bold=True, color="F59E0B")
    ws.merge_cells('G4:I4'); ws['G4'] = f"Total Revenue: Rs. {revenue:,.0f}"; ws['G4'].font = Font(bold=True, color="7C3AED")

    # Headers
    headers = ['#', 'Order ID', 'Customer', 'Email', 'Event', 'Date', 'Guests', 'Per Plate', 'Grand Total', 'Status', 'Created']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data rows
    for i, o in enumerate(orders, 1):
        row = 6 + i
        vals = [
            i, o.order_id,
            o.customer.name if o.customer else '',
            o.customer.email if o.customer else '',
            o.event_type or '', str(o.event_date) if o.event_date else '',
            o.guest_count, f"Rs. {o.per_plate:,.0f}",
            f"Rs. {o.grand_total:,.0f}", o.status.upper(),
            o.created_at.strftime('%d %b %Y %H:%M') if o.created_at else ''
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [1,7,10] else 'left')
            if col == 10:  # Status column
                cell.fill = status_fills.get(o.status, PatternFill())
                cell.font = Font(bold=True, size=10)

    # Column widths
    widths = [5, 14, 20, 25, 14, 14, 10, 14, 16, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Daily breakdown sheet
    if not date_str:
        ws2 = wb.create_sheet("Daily Summary")
        ws2.merge_cells('A1:E1')
        ws2['A1'] = "Daily Order Summary"
        ws2['A1'].font = title_font

        daily_headers = ['Date', 'Total Orders', 'Confirmed', 'Pending', 'Revenue']
        for col, h in enumerate(daily_headers, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')

        # Group by date
        daily = {}
        for o in orders:
            d = str(o.event_date) if o.event_date else 'No Date'
            daily.setdefault(d, {'total': 0, 'confirmed': 0, 'pending': 0, 'revenue': 0})
            daily[d]['total'] += 1
            if o.status == 'confirmed': daily[d]['confirmed'] += 1; daily[d]['revenue'] += o.grand_total
            if o.status == 'pending': daily[d]['pending'] += 1

        for i, (d, data) in enumerate(sorted(daily.items()), 1):
            row = 3 + i
            ws2.cell(row=row, column=1, value=d).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=2, value=data['total']).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=3, value=data['confirmed']).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=4, value=data['pending']).alignment = Alignment(horizontal='center')
            ws2.cell(row=row, column=5, value=f"Rs. {data['revenue']:,.0f}")

        for i, w in enumerate([14, 14, 14, 14, 18], 1):
            ws2.column_dimensions[chr(64+i)].width = w

    if is_master:
        fname = "MenuMatrix_Master_Orders.xlsx"
        # Save to exports/ dir only — never to project root (avoids triggering Flask reloader)
        path = os.path.join(EXCEL_DIR, fname)
    else:
        fname = f"orders_{date_str or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = os.path.join(EXCEL_DIR, fname)
    
    wb.save(path)
    return path

def update_master_excel():
    try:
        from app import Order
        orders = Order.query.order_by(Order.created_at.desc()).all()
        generate_orders_excel(orders, is_master=True)
    except Exception as e:
        print(f"Error updating master excel: {e}")
